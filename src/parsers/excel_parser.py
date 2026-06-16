from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

from src.core.exceptions import WorkbookError

# ExcelParser is a simple parser for .xlsx files that extracts the input and output sheet names for each resident.
# It is used to extract the input and output sheet names for each resident, which are then processed by the LLM.


@dataclass(frozen=True)
class WorkbookResidentSheet:
    input_sheet: str
    output_sheet: str
    resident_name: str


class ExcelParser:
    def __init__(self, path: Path) -> None:
        self.path = path

    def load(self):
        if not self.path.exists():
            raise WorkbookError(f"Workbook not found: {self.path}")
        return load_workbook(self.path)

    def resident_pairs(self) -> list[WorkbookResidentSheet]:
        workbook = self.load()
        pairs: list[WorkbookResidentSheet] = []
        for sheet_name in workbook.sheetnames:
            if not sheet_name.endswith(" - Input"):
                continue
            resident_name = sheet_name[: -len(" - Input")]
            output_sheet = f"{resident_name} - Your Output"
            if output_sheet not in workbook.sheetnames:
                raise WorkbookError(f"Missing output sheet for {resident_name}")
            pairs.append(WorkbookResidentSheet(sheet_name, output_sheet, resident_name))
        return pairs
