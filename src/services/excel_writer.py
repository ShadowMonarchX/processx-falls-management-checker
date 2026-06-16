from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from src.core.exceptions import WorkbookError
from src.core.models import ComplianceFlagModel
from src.utils.file_utils import ensure_parent


class ExcelWriter:
    def __init__(self, output_path: Path) -> None:
        self.output_path = output_path

    def open_workbook(self, source_path: Path):
        if not source_path.exists():
            raise WorkbookError(f"Workbook not found: {source_path}")
        return load_workbook(source_path)

    def write_flags(self, workbook, sheet_name: str, flags: list[ComplianceFlagModel]) -> None:
        if sheet_name not in workbook.sheetnames:
            raise WorkbookError(f"Missing output sheet: {sheet_name}")
        worksheet: Worksheet = workbook[sheet_name]
        start_row = self._next_empty_row(worksheet)
        for index, flag in enumerate(flags, start=start_row):
            worksheet[f"A{index}"] = flag.day_label
            worksheet[f"B{index}"] = flag.severity.value
            worksheet[f"C{index}"] = flag.field
            worksheet[f"D{index}"] = self._build_explanation(flag)

    def save(self, workbook) -> None:
        ensure_parent(self.output_path)
        workbook.save(self.output_path)

    def _next_empty_row(self, worksheet: Worksheet) -> int:
        row = 4
        while any(worksheet[f"{column}{row}"].value not in (None, "") for column in "ABCD"):
            row += 1
        return row

    def _build_explanation(self, flag: ComplianceFlagModel) -> str:
        return (
            f"Policy Rule: {flag.policy_rule_id}. "
            f"Trigger Condition: {flag.trigger_condition}. "
            f"Evidence Found: {flag.evidence_found[:220]}. "
            f"Missing Requirement: {flag.missing_requirement}. "
            f"Recommendation: {flag.recommendation}"
        )
