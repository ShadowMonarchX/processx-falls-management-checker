import logging
from pathlib import Path

from openpyxl.worksheet.worksheet import Worksheet

from src.ai.llm_client import LLMClient
from src.ai.prompt_builder import PromptBuilder
from src.core.exceptions import WorkbookError
from src.core.models import ComplianceFlagModel, ResidentNoteModel
from src.services.compliance_engine import ComplianceEngine
from src.services.excel_writer import ExcelWriter


# ValidationEngine is the main service that orchestrates the entire validation process by reading the resident notes from the input sheets, analyzing them using the ComplianceEngine, and writing the compliance flags into the output sheets for each resident.
# It is used to read the resident notes from the input sheets, analyze them using the ComplianceEngine, and write the compliance flags into the output sheets for each resident, which can then be reviewed by the care team to identify any documentation gaps and take corrective actions.
# The engine processes each resident's notes in chronological order (day-by-day) to ensure that the compliance flags are generated in the same order as the notes, and includes error handling to ensure that the output sheet exists for each resident, raising a WorkbookError if any sheet is missing. This helps ensure that the validation process runs smoothly and that the compliance flags are recorded accurately for review by the care team.
class ValidationEngine:
    def __init__(
        self,
        source_workbook: Path,
        output_workbook: Path,
        compliance_engine: ComplianceEngine,
    ) -> None:
        self.source_workbook = source_workbook
        self.output_workbook = output_workbook
        self.compliance_engine = compliance_engine
        self.prompt_builder = PromptBuilder()
        self.llm_client = LLMClient()
        self.logger = logging.getLogger("processx")

    def run(self) -> None:
        writer = ExcelWriter(self.output_workbook)
        workbook = writer.open_workbook(self.source_workbook)
        for sheet_name in workbook.sheetnames:
            if not sheet_name.endswith(" - Input"):
                continue
            resident_name = sheet_name[: -len(" - Input")]
            output_sheet = f"{resident_name} - Your Output"
            if output_sheet not in workbook.sheetnames:
                raise WorkbookError(
                    f"Output sheet missing for resident: {resident_name}"
                )
            flags = self._analyze_resident_sheet(workbook[sheet_name], resident_name)
            writer.write_flags(workbook, output_sheet, flags)
        writer.save(workbook)

    def _analyze_resident_sheet(
        self, worksheet: Worksheet, resident_name: str
    ) -> list[ComplianceFlagModel]:
        # The workbook is read row-by-row because each row represents a single daily note and
        # the output workbook expects one or more flags to be appended in the same order.
        flags: list[ComplianceFlagModel] = []
        for row in range(4, worksheet.max_row + 1):
            day = worksheet[f"A{row}"].value
            date = worksheet[f"B{row}"].value
            documented_by = worksheet[f"C{row}"].value
            note_text = worksheet[f"D{row}"].value
            if not day or not note_text:
                continue
            note = ResidentNoteModel(
                resident_name=resident_name,
                day_label=str(day),
                date=str(date or ""),
                documented_by=str(documented_by or ""),
                note_text=str(note_text or ""),
            )
            prompt = self.prompt_builder.build_extraction_prompt(
                resident_name=resident_name,
                day=str(day),
                note_text=note.note_text,
                rules=self.compliance_engine.rules,
            )
            fallback_payload = {
                "resident_name": resident_name,
                "day": str(day),
                "observations": {
                    "documented_by": note.documented_by,
                    "date": note.date,
                },
                "evidence": [note.note_text[:500]],
                "missing_items": [],
                "confidence": 0.2,
                "provider": "fallback",
                "model": "rule-derived",
            }
            llm_result = self.llm_client.extract(prompt, fallback_payload)
            self.logger.info(
                "ai_extraction",
                extra={
                    "provider": llm_result.provider,
                    "model": llm_result.model,
                    "latency_ms": llm_result.latency_ms,
                    "resident": resident_name,
                    "day": str(day),
                    "confidence": llm_result.payload.confidence,
                },
            )
            flags.extend(self.compliance_engine.analyze(note, llm_result.payload))
        return flags
