from pathlib import Path

from openpyxl import load_workbook

from src.core.constants import COMPLETED_WORKBOOK_PATH, INPUT_WORKBOOK_PATH, POLICY_PATH
from src.parsers.policy_parser import PolicyParserImpl
from src.services.compliance_engine import ComplianceEngine
from src.services.validation_engine import ValidationEngine


def test_end_to_end_output_creation(tmp_path):
    output_path = tmp_path / "completed_output.xlsx"
    rules = PolicyParserImpl(POLICY_PATH).parse()
    engine = ComplianceEngine(rules)
    validator = ValidationEngine(INPUT_WORKBOOK_PATH, output_path, engine)
    validator.run()
    wb = load_workbook(output_path)
    assert "Alice Nguyen - Your Output" in wb.sheetnames
    assert wb["Alice Nguyen - Your Output"]["A4"].value == ""

